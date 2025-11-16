import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime

# ✅ Import models explicitly from their respective apps
from certificates.models import CertificateRequest, BusinessPermit
from blotter.models import BlotterReport
from complaints.models import Complaint
from emergency.models import EmergencyReport


def export_all_reports_excel(request):
    # Create a workbook and remove the default sheet
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def add_sheet(title, headers, rows):
        ws = wb.create_sheet(title)
        ws.append(headers)

        for row in rows:
            clean_row = []
            for value in row:
                # ✅ Remove timezone info from datetime fields
                if isinstance(value, datetime):
                    if timezone.is_aware(value):
                        value = timezone.make_naive(value)
                clean_row.append(value)
            ws.append(clean_row)

        # ✅ Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    # ✅ Certificates
    certificate_qs = CertificateRequest.objects.all().values_list(
        "request_number",
        "certificate_type",
        "first_name",
        "last_name",
        "middle_name",
        "complete_address",
        "contact_number",
        "email_address",
        "purpose",
        "business_name",
        "status",
        "created_at",
    )
    add_sheet(
        "Certificates",
        [
            "Request Number",
            "Certificate Type",
            "First Name",
            "Last Name",
            "Middle Name",
            "Complete Address",
            "Contact Number",
            "Email",
            "Purpose",
            "Business Name",
            "Status",
            "Created At",
        ],
        certificate_qs,
    )

    # ✅ Business Permits
    permit_qs = BusinessPermit.objects.all().values_list(
        "business_name",
        "business_type",
        "owner_name",
        "business_address",
        "contact_number",
        "owner_address",
        "business_description",
        "is_renewal",
        "status",
        "created_at",
    )
    add_sheet(
        "Business Permits",
        [
            "Business Name",
            "Business Type",
            "Owner Name",
            "Business Address",
            "Contact Number",
            "Owner Address",
            "Description",
            "Renewal",
            "Status",
            "Created At",
        ],
        permit_qs,
    )

    # ✅ Complaints
    complaint_qs = Complaint.objects.all().values_list(
        "reference_number",
        "type",
        "fullname",
        "contact_number",
        "address",
        "email_address",
        "subject",
        "detailed_description",
        "respondent_name",
        "respondent_address",
        "latitude",
        "longitude",
        "status",
        "priority",
        "date_filed",
    )
    add_sheet(
        "Complaints",
        [
            "Reference Number",
            "Type",
            "Full Name",
            "Contact Number",
            "Address",
            "Email",
            "Subject",
            "Description",
            "Respondent Name",
            "Respondent Address",
            "Latitude",
            "Longitude",
            "Status",
            "Priority",
            "Date Filed",
        ],
        complaint_qs,
    )

    # ✅ Emergency Reports
    emergency_qs = EmergencyReport.objects.all().values_list(
        "name",
        "incident_type",
        "description",
        "latitude",
        "longitude",
        "alert_message",
        "status",
        "location_text",
        "contact_number",
        "submitted_at",
    )
    add_sheet(
        "Emergencies",
        [
            "Name",
            "Incident Type",
            "Description",
            "Latitude",
            "Longitude",
            "Alert Message",
            "Status",
            "Location",
            "Contact Number",
            "Submitted At",
        ],
        emergency_qs,
    )

    # ✅ Blotter Reports
    blotter_qs = BlotterReport.objects.all().values_list(
        "report_number",
        "complainant_name",
        "respondent_name",
        "incident_type",
        "incident_date",
        "incident_time",
        "location",
        "description",
        "priority",
        "status",
        "resolution_notes",
        "created_at",
    )
    add_sheet(
        "Blotter Reports",
        [
            "Report #",
            "Complainant",
            "Respondent",
            "Incident Type",
            "Incident Date",
            "Incident Time",
            "Location",
            "Description",
            "Priority",
            "Status",
            "Resolution Notes",
            "Created At",
        ],
        blotter_qs,
    )

    # ✅ Prepare Excel file response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Consolidated_Reports_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response
